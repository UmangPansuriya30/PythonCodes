from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re

def save_data(filename, row_data, sheet_name="Sheet1"):
    from openpyxl import Workbook, load_workbook
    import os
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(title=sheet_name)
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(["GalleryName", "Location", "Website", "Email", "Contact", "Source"])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
    sheet.append(row_data)
    wb.save(filename)
try:
    driver = webdriver.Chrome()
    driver.get("https://www.artspace.com/partners")

    WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.XPATH, "//button[@id='hs-eu-confirmation-button']"))
    )
    button = driver.find_element(By.XPATH, "//button[@id='hs-eu-confirmation-button']")
    button.click()
    flag= True
    for i in range(1,3):
        driver.switch_to.frame(driver.find_element(By.XPATH, f"(//iframe[@title='Popup CTA'])[{i}]"))
        button2 = driver.find_element(By.XPATH, "//div[@id='interactive-close-button']")
        button2.click()
        driver.switch_to.parent_frame()

    SCROLL_PAUSE = 2

    while True:
        last_height = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    elements = driver.find_elements(By.XPATH, "//div[contains(@class,'element')]/a")
    print(f"Total galleries found: {len(elements)}")
    Links = [ele.get_attribute("href") for ele in elements]
    for link in Links:
        driver.get(link)
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//img[@itemprop='image']"))
        )
        Title = re.match(r".*(?=\sFollow)", nodes[0].text).group(0) if (nodes := driver.find_elements(By.XPATH, "//div[@itemprop='description']/h1")) else ""
        element = nodes[0] if (nodes := driver.find_elements(By.XPATH, "//address")) else ""
        if element.text:
            # fieldNodes = element.find_elements(By.XPATH, ".//div[@itemprop='location']")
            City = City = nodes[0].text if (nodes := element.find_elements(By.XPATH, ".//div[@itemprop='location']")) else ""
            Address = nodes[0].text if (nodes := element.find_elements(By.XPATH, ".//div[@itemprop='address']")) else "" 
            Phone = nodes[0].text if (nodes := element.find_elements(By.XPATH, ".//div[@itemprop='telephone']")) else ""
        
        element = nodes[0] if (nodes := driver.find_elements(By.XPATH, "//ul[@class='partner-links']")) else ""
        if element.text:
            Website = nodes[0].get_attribute("href") if (nodes := element.find_elements(By.XPATH, ".//li[@class='partner-website']/a")) else ""
        
        save_data("C:\\WorkingDirectory\\Gallery3.xlsx", [Title.strip(), f"{Address.strip()}, {City.strip()}", Website.strip(), "", Phone.strip(), link.strip()],"Gallery-3")
        print([Title.strip(), f"{Address.strip()}, {City.strip()}", Website.strip(), "", Phone.strip(), link.strip()])
    print("--------------------Done--------------------")
except Exception as e:
    print("Element not found:", e)  
finally:
    driver.quit()
